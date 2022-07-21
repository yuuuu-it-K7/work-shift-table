# インポート
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side

import calendar
import jpholiday



year_given = 2022

# ファイルの新規作成
wb = openpyxl.Workbook()
file_name = 'kintai.xlsx'
wb.save(file_name)


def create_title(ws):
    """タイトル行の作成"""
    # セルへ書き込む値を指定
    ws.cell(1, 1).value = '2022年'
    ws.cell(2, 5).value = '日'
    ws.cell(3, 5).value = '曜日'
    ws.cell(4, 5).value = '予算'
    ws.cell(5, 5).value = '前年'
    ws.cell(6, 3).value = '2021年度客数'
    ws.cell(7, 3).value = '2021年度前年'
    ws.cell(8, 3).value = '2020年度客数'
    # フォントスタイルの指定（月・日・曜日）
    ws.cell(2, 5).font = Font(name='ヒラギノ明朝 Pro', size=18)
    ws.cell(3, 5).font = Font(name='ヒラギノ明朝 Pro', size=18)
    ws.cell(4, 5).font = Font(name='ヒラギノ明朝 Pro', size=18)
    ws.cell(5, 5).font = Font(name='ヒラギノ明朝 Pro', size=18)
    ws.cell(6, 3).font = Font(name='ヒラギノ明朝 Pro', size=18)
    ws.cell(7, 3).font = Font(name='ヒラギノ明朝 Pro', size=18)
    ws.cell(8, 3).font = Font(name='ヒラギノ明朝 Pro', size=18)
    # セル結合の指定
    ws.merge_cells('C2:D5')
    ws.merge_cells('C6:E6') 
    ws.merge_cells('C7:E7') 
    ws.merge_cells('C8:E8') 
    # セル中央揃えの指定
    for row in ws["C2:AJ8"]:
        for cell in row:
            cell.alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = False)
    # セル枠線の指定
    side1 = Side(style='thin', color='000000')
    side2 = Side(style='thick', color='000000')
    border1 = Border(top=side1, bottom=side1, left=side1, right=side1)
    border2 = Border(top=side1, bottom=side1, left=side1, right=side1)
    for row in ws['C2:AJ8']:
        for cell in row:
            cell.border = border1
    for row in ws['C2:D5']:
        for cell in row:
            cell.border = border2 
    # セル幅の指定（月・日・曜日）
    ws.column_dimensions[get_column_letter(1)].width = 5
    ws.column_dimensions[get_column_letter(2)].width = 5
    # フォントスタイル、セルの色、セル幅の指定（月・日・曜日　以外）
    for i in range(6, 37):
        ws.cell(2, i).font = Font(name='ヒラギノ明朝 Pro', size=18, bold=True, italic=False)
        ws.cell(3, i).font = Font(name='ヒラギノ明朝 Pro', size=18, bold=True, italic=False)


for month_given in range(1, 13):
    sheet_name = '{}年{}月'.format(year_given, month_given)
    ws = wb.create_sheet(sheet_name)
    c = calendar.Calendar(firstweekday=0)
    monthdatescalendar = c.monthdatescalendar(year_given, month_given)
    # その月の全日付を取得（前月と次月の日付は除く）
    date_list = []
    for dates in monthdatescalendar:
        for date in dates:
            if date.month == month_given:
                date_list.append(date)
    create_title(ws)
    for row, i in enumerate(date_list):
        row += 6
        date = i.day
        day = i.strftime('%a')
        day = day.replace('Sun', '日').replace('Mon', '月').replace('Tue', '火').replace('Wed', '水')\
        .replace('Thu', '木').replace('Fri', '金').replace('Sat', '土') #英語表記を日本語表記に変換
        holiday_name = jpholiday.is_holiday_name(i)
        ws.cell(2, row).value = date
        if holiday_name:
            holiday_name = day
            ws.cell(2, row).fill = PatternFill(patternType='solid', fgColor='ffff33')
            ws.cell(3, row).fill = PatternFill(patternType='solid', fgColor='ffff33')
        else:
            if day == '土':
                ws.cell(2, row).fill = PatternFill(patternType='solid', fgColor='ffff33')
                ws.cell(3, row).fill = PatternFill(patternType='solid', fgColor='ffff33')
            elif day == '日':
                ws.cell(2, row).fill = PatternFill(patternType='solid', fgColor='ffff33')
                ws.cell(3, row).fill = PatternFill(patternType='solid', fgColor='ffff33')
        ws.cell(3, row).value = day


        

wb.save(file_name) # ファイルを保存
